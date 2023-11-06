from datetime import date
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from docxtpl import DocxTemplate
from math import degrees
import pandas as pd

MAX_ROWS_LIMIT = 100000


def get_workbook(file_path: str, comm_signal) -> Workbook:
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        comm_signal.emit('c', f'Fajl {file_path} nije pronađen, proverite putanju.')
        return 'Error'
    except InvalidFileException:
        comm_signal.emit('c', 'Pogrešan tip fajla. Moguće je učitati samo *.xlsx fajlove.')
        return 'Error'
    except PermissionError:
        comm_signal.emit('c', 'Zatvorite ulazni fajl sa podacima.')
        return 'Error'

    if wb:
        return wb
    else:
        comm_signal.emit('c', 'Morate izabrati Excel fajl pre pokretanja!')
        return 'Error'


def extract_excel_data(workbook: Workbook, start_cell: str, worksheet, comm_signal,
                       extraction_type='coordinate') -> List[list]:

    ws = workbook.get_sheet_by_name(worksheet)

    if extraction_type in ['coordinate', 'distances']:
        col_span = 3
    else:
        col_span = 5

    try:
        col_start = start_cell.rstrip('0123456789')
        row_start = int(start_cell[len(col_start):])
    except ValueError:
        comm_signal.emit('c', "Greška prilikom određivanja početne čelije. Format mora biti: '<slovo><broj>'")
        return 'Error'

    col_start_index = column_index_from_string(col_start)

    row_data = []

    while row_start < MAX_ROWS_LIMIT:
        col_data = []
        for col in range(col_start_index, col_start_index + col_span):
            index = get_column_letter(col) + str(row_start)
            if ws[index].value != 'None':
                col_data.append(ws[index].value)

        none_row_check = [1 if value is None else 0 for value in col_data]

        if sum(none_row_check) == 0:
            row_data.append(col_data)
            row_start += 1
        elif sum(none_row_check) < len(none_row_check):
            comm_signal.emit('w', f'Postoji kolona/kolone bez vrednosti!\nProverite red {row_start} '
                                  f'u fajlu sa ulaznim podacima i pokrenite ponovo.')
            return 'Error'
        else:
            # case where all column values in row span are None, no message needed - end of data condition
            break

    return row_data


def excel_export(data: dict, file_path, comm_signal, points):

    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            row = 1
            col = 1
            for k, v in data.items():
                if k == 'zenitni uglovi':
                    v = [[key, value] for key, value in v.items()]

                if k == 'standardne devijacije':
                    v = [(point.id, point.sigma_y, point.sigma_x, point.sigma_p) for point in points]
                    header = pd.DataFrame([['Tačka', '\u03C3Y', '\u03C3X', '\u03C3P']])
                    header.to_excel(writer, sheet_name=k, index=False, header=False, startrow=row, startcol=col)
                    row += 1

                if k == 'elipse gresaka':
                    v = []
                    for point in points:
                        v.append([point.id, point.ellipse.theta, point.ellipse.a, point.ellipse.b])
                    header = pd.DataFrame([['Tačka', '\u03B8', 'A', 'B']])
                    header.to_excel(writer, sheet_name=k, index=False, header=False, startrow=row, startcol=col)
                    row += 1

                if k == 'data snooping':
                    start_col = 1
                    start_row = 1
                    for iteration in v:
                        for key, value in iteration.items():
                            if key == 'iteracija':
                                export = pd.DataFrame([f'{key}: {value}'])
                                export.to_excel(writer, sheet_name=k, index=False, header=False, startrow=start_row,
                                                startcol=start_col)
                                start_row += 2
                            elif key in ['m0', 'T', 'F', 'Norm', 'izbaceno merenje']:
                                export = pd.DataFrame([key, value])
                                export.to_excel(writer, sheet_name=k, index=False, header=False, startrow=start_row,
                                                startcol=start_col)
                                start_row += 2
                            else:
                                header = pd.DataFrame([key])
                                header.to_excel(writer, sheet_name=k, index=False, header=False, startrow=start_row,
                                                startcol=start_col)
                                start_row += 2

                                export = pd.DataFrame(value)
                                export.to_excel(writer, sheet_name=k, index=False, header=False, startrow=start_row,
                                                startcol=start_col)
                                start_row += value.shape[0] + 3
                    continue

                if k == 'pouzdanost':
                    ri, gi = v
                    v = [[i+1, r, g] for i, (r, g) in enumerate(zip(ri, gi))]

                    header = pd.DataFrame([['Indeks', 'ri', 'Gi']])
                    header.to_excel(writer, sheet_name=k, index=False, header=False, startrow=row, startcol=col)
                    row += 1

                export = pd.DataFrame(v)
                export.to_excel(writer, sheet_name=k, index=False, header=False, startrow=row, startcol=col)

            header = pd.DataFrame([['Tačka', 'Y', 'X', 'Z0 [rad]']])
            header.to_excel(writer, sheet_name='izravnate koordinate', index=False, header=False, startrow=row,
                            startcol=col)
            corrected_points = []
            for point in points:
                corrected_points.append([point.id, point.y, point.x, point.z0])
            corrected_points = pd.DataFrame(corrected_points)
            corrected_points.to_excel(writer, sheet_name='izravnate koordinate', index=False, header=False,
                                      startrow=row + 1, startcol=col)

    except PermissionError:
        comm_signal.emit('c', f"Excel fajl na putanji:\n'{file_path}'\n je otvoren u Excel-u. Zatvorite postojeći "
                              f"fajl ili kreirajte novi.")
        return 'Error'


def word_report(file_path, old_points, new_points, directions, distances,
                dropped_msm, defect, station_num, params, comm_signal, excel_export):
    doc = DocxTemplate('report_template.docx')
    unknown_params_num = 2 * len(old_points) + station_num
    if 'Sve tačke' in params['datumCoords']:
        point_id = [point.id for point in old_points]
        coords = ', '.join(point_id)
    else:
        coords = ', '.join(params['datumCoords'])

    if params['datumMethod'] == 'min_trace':
        datum_def = f'минималним трагом на тачке: {coords}'
    else:
        point, coord = params['datumCoords'].split(';')
        coord = coord.split('-')
        datum_def = f'класичном методом, фиксирањем координата тачке {point} и коодинате {coord[0]} тачке {coord[1]}'

    points_table = []
    [points_table.append({'id': point.id,
                          'y': round(point.y, 4),
                          'x': round(point.x, 4)}) for point in old_points]

    p_diag = excel_export['P'].diagonal()
    f = excel_export['f']
    v = excel_export['v']
    qv_diag = excel_export['qv'].diagonal()
    ql_diag = excel_export['ql'].diagonal()
    ri, gi = excel_export['pouzdanost']

    dir_table = []
    [dir_table.append({'from': dire.from_,
                       'to': dire.to,
                       'value': round(degrees(dire.value), 5),
                       'p': round(p_diag[i], 2),
                       'f': round(f[i], 2),
                       'v': round(v[i], 2),
                       'r': round(ri[i], 2),
                       'g': round(gi[i], 2),
                       'qv': round(qv_diag[i], 2),
                       'ql': round(ql_diag[i], 2)}) for i, dire in enumerate(directions)]

    dis_table = []
    dir_len = len(directions)
    [dis_table.append({'from': dist.from_,
                       'to': dist.to,
                       'value': round(dist.value, 4),
                       'p': round(p_diag[i + dir_len], 2),
                       'f': round(f[i + dir_len], 2),
                       'v': round(v[i + dir_len], 2),
                       'r': round(ri[i + dir_len], 2),
                       'g': round(gi[i + dir_len], 2),
                       'qv': round(qv_diag[i + dir_len], 2),
                       'ql': round(ql_diag[i + dir_len], 2)}) for i, dist in enumerate(distances)]

    dropped_msm_table = []
    if dropped_msm is not None:
        for msm_type, msm, t, f, w, n in dropped_msm:
            if msm_type == 'Pravac':
                new_type = 'Правац'
            else:
                new_type = 'Дужина'

            dropped_msm_table.append({'type': new_type,
                                      'from': msm.from_,
                                      'to': msm.to,
                                      'value': round(msm.value, 5),
                                      't': round(t, 3),
                                      'f': round(f, 3),
                                      'w': round(w, 2),
                                      'n': round(n, 3)})

    stdev_table = []
    [stdev_table.append({'id': point.id,
                         'y': round(point.sigma_y, 2),
                         'x': round(point.sigma_x, 2),
                         'p': round(point.sigma_p, 2)}) for point in new_points]

    ellipse_table = []
    [ellipse_table.append({'id': point.id,
                           'a': round(point.ellipse.a, 3),
                           'b': round(point.ellipse.b, 3),
                           'deg': point.ellipse.thetaDMS[0],
                           'min': point.ellipse.thetaDMS[1],
                           'sec': point.ellipse.thetaDMS[2]}) for point in new_points]

    corrected_points_table = []
    [corrected_points_table.append({'id': point.id,
                                    'y': round(point.y, 5),
                                    'x': round(point.x, 5)}) for point in new_points]

    context = {'today': date.today().strftime("%d.%m.%Y."),
               'point_num': len(old_points),
               'station_num': station_num,
               'direction_num': len(directions),
               'distance_num': len(distances),
               'msm_num': len(directions) + len(distances),
               'unknown_parameters_num': unknown_params_num,
               'degrees_of_freedom': len(directions) + len(distances) - unknown_params_num + defect,
               'sigma0': params['sigma0'],
               'datum_def': datum_def,
               'points_table': points_table,
               'dir_table': dir_table,
               'dis_table': dis_table,
               'alpha': round(float(params['alphaCoef']), 2),
               'dropped_msm_table': dropped_msm_table,
               'stdev_table': stdev_table,
               'ellipse_table': ellipse_table,
               'corrected_points_table': corrected_points_table
               }
    doc.render(context)
    try:
        doc.save(file_path)
    except PermissionError:
        comm_signal.emit('c', f"Word fajl na putanji:\n'{file_path}'\n je otvoren. Zatvorite postojeći fajl "
                              f"ili kreirajte novi.")
        return 'Error'


def point_csv(points, file_path, comm_signal):
    if file_path is None:
        return

    point_csv = [point.toCSV() for point in points]
    try:
        file = open(file_path, 'w')
    except PermissionError:
        comm_signal.emit('c', f"CSV fajl na putanji:\n'{file_path}'\n je otvoren. Zatvorite postojeći fajl "
                              f"ili kreirajte novi.")
        return 'Error'

    for line in point_csv:
        file.write(line)
    file.close()


def scr_file(data, file_path, comm_signal):
    if file_path is None:
        return

    try:
        file = open(file_path, 'w')
    except PermissionError:
        comm_signal.emit('c', f"CAD Script fajl na putanji:\n'{file_path}'\n je otvoren. Zatvorite postojeći fajl "
                              f"ili kreirajte novi.")
        return 'Error'

    for line in data:
        file.write(line)
        file.write('\n')
    file.close()
